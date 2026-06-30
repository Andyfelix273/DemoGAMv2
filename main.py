"""
GIS Asset Manager - Backend FastAPI
Versione: 2.0
Autore: Felix / KeyBiz

Endpoint:
  POST /auth/login              - login, restituisce JWT
  GET  /auth/me                 - profilo utente corrente

  GET  /api/assets              - lista asset in GeoJSON
  GET  /api/assets/{id}         - dettaglio asset + monitoraggio
  POST /api/assets              - crea asset (solo admin)
  PUT  /api/assets/{id}         - modifica asset (solo admin)
  DELETE /api/assets/{id}       - elimina asset (solo admin)

  GET  /api/stats               - KPI aggregati per dashboard
  GET  /api/thresholds          - lista soglie allarme
  PUT  /api/thresholds/{id}     - modifica soglia (solo admin)
  GET  /api/alarms              - lista allarmi attivi
  POST /api/alarms/{id}/ack     - acknowledge allarme (solo admin)

  GET  /api/config              - configurazione applicazione
  PUT  /api/config              - aggiorna configurazione (solo admin)

  POST /api/assets/import       - import da Excel (solo admin)
"""

import os
import sqlite3
import io
import asyncio
import random
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse
from jose import JWTError, jwt
from passlib.context import CryptContext
from pydantic import BaseModel

# ── Configurazione ──────────────────────────────────────────────────────────
SECRET_KEY  = "gis-asset-manager-secret-key-2026"
ALGORITHM   = "HS256"
TOKEN_EXPIRE_MINUTES = 480  # 8 ore

DB_PATH     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets.db")

# ── Controllo Cartelle Statiche ──────────────────────────────────────────────
# Assicura che la struttura delle cartelle statiche esista al primo avvio
STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
SUBDIRS = ["css", "js", "img", "data", "pages"]

for sd in SUBDIRS:
    path = os.path.join(STATIC_DIR, sd)
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)
        print(f"[setup] creata cartella: {path}")


pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")

app = FastAPI(title="GIS Asset Manager API", version="2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Utilità database ─────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()


# ── Modelli Pydantic ─────────────────────────────────────────────────────────
class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    role: str
    nome_completo: Optional[str] = None

class AssetCreate(BaseModel):
    codice: str
    nome: str
    tipo: str
    indirizzo: Optional[str] = None
    citta: Optional[str] = None
    provincia: Optional[str] = None
    cap: Optional[str] = None
    lat: float
    lon: float
    referente: Optional[str] = None
    telefono: Optional[str] = None
    email: Optional[str] = None
    superficie_mq: Optional[int] = None
    anno_costruzione: Optional[int] = None
    stato: Optional[str] = "attivo"
    note: Optional[str] = None

class AssetUpdate(BaseModel):
    nome: Optional[str] = None
    tipo: Optional[str] = None
    indirizzo: Optional[str] = None
    citta: Optional[str] = None
    provincia: Optional[str] = None
    cap: Optional[str] = None
    lat: Optional[float] = None
    lon: Optional[float] = None
    referente: Optional[str] = None
    telefono: Optional[str] = None
    email: Optional[str] = None
    superficie_mq: Optional[int] = None
    anno_costruzione: Optional[int] = None
    stato: Optional[str] = None
    note: Optional[str] = None

class ThresholdUpdate(BaseModel):
    warning_value: float
    alarm_value: float
    is_percentage: Optional[int] = None   # 1 = soglie in %, 0 = valore assoluto
    campo_totale: Optional[str] = None    # campo denominatore per il calcolo %
    inverso: Optional[int] = None         # 1 = basso è brutto (es. mezzi disponibili)

class ConfigUpdate(BaseModel):
    tema: Optional[str] = None
    lingua: Optional[str] = None
    rotte_marittime: Optional[str] = None


# ── Autenticazione JWT ───────────────────────────────────────────────────────
def crea_token(data: dict) -> str:
    payload = data.copy()
    payload["exp"] = datetime.now(timezone.utc) + timedelta(minutes=TOKEN_EXPIRE_MINUTES)
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def get_utente_corrente(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        role     = payload.get("role")
        if not username:
            raise HTTPException(status_code=401, detail="Token non valido")
        return {"username": username, "role": role}
    except JWTError:
        raise HTTPException(status_code=401, detail="Token non valido o scaduto")


def richiedi_admin(utente=Depends(get_utente_corrente)):
    if utente["role"] != "admin":
        raise HTTPException(status_code=403, detail="Accesso riservato agli amministratori")
    return utente


# ── Endpoint autenticazione ──────────────────────────────────────────────────
@app.post("/auth/login", response_model=TokenResponse)
def login(form: OAuth2PasswordRequestForm = Depends(), db=Depends(get_db)):
    utente = db.execute(
        "SELECT * FROM users WHERE username=?", (form.username,)
    ).fetchone()
    if not utente or not pwd_context.verify(form.password, utente["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenziali non valide")
    token = crea_token({"sub": utente["username"], "role": utente["role"]})
    return {
        "access_token": token,
        "token_type": "bearer",
        "role": utente["role"],
        "nome_completo": utente["nome_completo"]
    }


@app.get("/auth/me")
def profilo(utente=Depends(get_utente_corrente), db=Depends(get_db)):
    row = db.execute(
        "SELECT username, role, nome_completo, email FROM users WHERE username=?",
        (utente["username"],)
    ).fetchone()
    return dict(row)


# ── Endpoint asset ───────────────────────────────────────────────────────────
@app.get("/api/assets")
def lista_assets(db=Depends(get_db), _=Depends(get_utente_corrente)):
    """Restituisce tutti gli asset in formato GeoJSON FeatureCollection."""
    rows = db.execute("SELECT * FROM assets ORDER BY codice").fetchall()
    features = []
    for r in rows:
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [r["lon"], r["lat"]]},
            "properties": {k: r[k] for k in r.keys() if k not in ("lat", "lon")}
        })
    return {"type": "FeatureCollection", "features": features}


@app.get("/api/assets/{asset_id}")
def dettaglio_asset(asset_id: int, db=Depends(get_db), _=Depends(get_utente_corrente)):
    """Restituisce il dettaglio di un asset con i dati di monitoraggio."""
    asset = db.execute("SELECT * FROM assets WHERE id=?", (asset_id,)).fetchone()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset non trovato")
    mon = db.execute(
        "SELECT * FROM monitoraggio WHERE asset_id=?", (asset_id,)
    ).fetchone()
    allarmi = db.execute(
        "SELECT * FROM alarms WHERE asset_id=? AND ack_at IS NULL ORDER BY created_at DESC",
        (asset_id,)
    ).fetchall()
    esg = db.execute(
        "SELECT * FROM esg WHERE asset_id=?", (asset_id,)
    ).fetchone()
    return {
        "asset": dict(asset),
        "monitoraggio": dict(mon) if mon else {},
        "allarmi": [dict(a) for a in allarmi],
        "esg": dict(esg) if esg else {}
    }


@app.post("/api/assets", status_code=201)
def crea_asset(payload: AssetCreate, db=Depends(get_db), _=Depends(richiedi_admin)):
    """Crea un nuovo asset (solo admin)."""
    try:
        cur = db.execute("""
            INSERT INTO assets
            (codice,nome,tipo,indirizzo,citta,provincia,cap,lat,lon,
             referente,telefono,email,superficie_mq,anno_costruzione,stato,note)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (payload.codice, payload.nome, payload.tipo, payload.indirizzo,
              payload.citta, payload.provincia, payload.cap, payload.lat, payload.lon,
              payload.referente, payload.telefono, payload.email,
              payload.superficie_mq, payload.anno_costruzione, payload.stato, payload.note))
        db.commit()
        return {"id": cur.lastrowid, "messaggio": "Asset creato"}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Codice asset già esistente")


@app.put("/api/assets/{asset_id}")
def modifica_asset(asset_id: int, payload: AssetUpdate, db=Depends(get_db), _=Depends(richiedi_admin)):
    """Modifica un asset esistente (solo admin)."""
    asset = db.execute("SELECT id FROM assets WHERE id=?", (asset_id,)).fetchone()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset non trovato")
    campi = {k: v for k, v in payload.dict().items() if v is not None}
    if not campi:
        raise HTTPException(status_code=400, detail="Nessun campo da aggiornare")
    set_clause = ", ".join(f"{k}=?" for k in campi)
    db.execute(f"UPDATE assets SET {set_clause} WHERE id=?", (*campi.values(), asset_id))
    db.commit()
    return {"messaggio": "Asset aggiornato"}


@app.delete("/api/assets/{asset_id}")
def elimina_asset(asset_id: int, db=Depends(get_db), _=Depends(richiedi_admin)):
    """Elimina un asset (solo admin)."""
    asset = db.execute("SELECT id FROM assets WHERE id=?", (asset_id,)).fetchone()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset non trovato")
    db.execute("DELETE FROM monitoraggio WHERE asset_id=?", (asset_id,))
    db.execute("DELETE FROM alarms WHERE asset_id=?", (asset_id,))
    db.execute("DELETE FROM assets WHERE id=?", (asset_id,))
    db.commit()
    return {"messaggio": "Asset eliminato"}


# ── Endpoint ESG aggregato ──────────────────────────────────────────────────
@app.get("/api/stats/esg")
def statistiche_esg(db=Depends(get_db), _=Depends(get_utente_corrente)):
    """Restituisce KPI ESG aggregati per la panoramica operativa."""
    row = db.execute("""
        SELECT
            SUM(e.co2_totale_kg_giorno)      AS co2_tot_kg,
            AVG(e.consumo_kwh_giorno)         AS kwh_medio,
            SUM(e.consumo_kwh_giorno)         AS kwh_tot,
            AVG(e.consumo_m3_acqua_giorno)    AS m3_medio,
            SUM(e.consumo_m3_acqua_giorno)    AS m3_tot,
            COUNT(CASE WHEN e.rating_esg='A' THEN 1 END) AS rating_a,
            COUNT(CASE WHEN e.rating_esg='B' THEN 1 END) AS rating_b,
            COUNT(CASE WHEN e.rating_esg='C' THEN 1 END) AS rating_c,
            COUNT(CASE WHEN e.rating_esg='D' THEN 1 END) AS rating_d,
            COUNT(*) AS n_asset
        FROM esg e
    """).fetchone()
    if not row or not row["n_asset"]:
        return {}
    co2_anno_t = round((row["co2_tot_kg"] or 0) * 365 / 1000, 1)
    # Rating flotta: media ponderata A=4, B=3, C=2, D=1
    punteggio = (row["rating_a"]*4 + row["rating_b"]*3 + row["rating_c"]*2 + row["rating_d"]*1) / row["n_asset"]
    if punteggio >= 3.5:
        rating_flotta = "A"
    elif punteggio >= 2.5:
        rating_flotta = "B"
    elif punteggio >= 1.5:
        rating_flotta = "C"
    else:
        rating_flotta = "D"
    return {
        "co2_totale_kg_giorno": round(row["co2_tot_kg"] or 0, 1),
        "co2_anno_tonnellate": co2_anno_t,
        "kwh_medio_giorno": round(row["kwh_medio"] or 0, 1),
        "kwh_totale_giorno": round(row["kwh_tot"] or 0, 1),
        "m3_medio_giorno": round(row["m3_medio"] or 0, 2),
        "m3_totale_giorno": round(row["m3_tot"] or 0, 1),
        "rating_flotta": rating_flotta,
        "distribuzione_rating": {
            "A": row["rating_a"], "B": row["rating_b"],
            "C": row["rating_c"], "D": row["rating_d"]
        }
    }


# ── Endpoint statistiche ─────────────────────────────────────────────────────
@app.get("/api/stats")
def statistiche(db=Depends(get_db), _=Depends(get_utente_corrente)):
    """Restituisce KPI aggregati per la dashboard di sintesi."""
    totale = db.execute("SELECT COUNT(*) FROM assets").fetchone()[0]
    per_tipo = {
        r["tipo"]: r["cnt"]
        for r in db.execute("SELECT tipo, COUNT(*) as cnt FROM assets GROUP BY tipo").fetchall()
    }

    # Legge soglie dal DB per i KPI principali (evita valori hardcoded)
    def get_soglie(tipo, campo):
        t = db.execute(
            "SELECT warning_value, alarm_value, inverso FROM thresholds WHERE asset_tipo=? AND campo=?",
            (tipo, campo)
        ).fetchone()
        if t:
            return {"soglia_warning": t["warning_value"], "soglia_alarm": t["alarm_value"],
                    "inverso": t["inverso"] or 0}
        return {"soglia_warning": 70.0, "soglia_alarm": 90.0, "inverso": 0}

    # KPI uffici: occupancy media (dipendenti/capienza * 100)
    uffici = db.execute("""
        SELECT AVG(CAST(m.dipendenti_presenti AS REAL)/m.capienza_massima*100) as occ
        FROM monitoraggio m JOIN assets a ON a.id=m.asset_id
        WHERE a.tipo='ufficio' AND m.capienza_massima > 0
    """).fetchone()
    s_uff = get_soglie('ufficio', 'dipendenti_presenti')
    occ_pct = round(uffici["occ"] or 0, 1)

    # KPI stabilimenti: linee attive / totali (percentuale)
    stab = db.execute("""
        SELECT SUM(m.linee_produzione_attive) as att, SUM(m.linee_produzione_totali) as tot
        FROM monitoraggio m JOIN assets a ON a.id=m.asset_id
        WHERE a.tipo='stabilimento'
    """).fetchone()
    s_stab = get_soglie('stabilimento', 'linee_produzione_attive')
    stab_att = int(stab["att"] or 0)
    stab_tot = int(stab["tot"] or 0)
    stab_pct = round((stab_att / stab_tot * 100) if stab_tot > 0 else 0, 1)

    # KPI magazzini: saturazione media (già %)
    mag = db.execute("""
        SELECT AVG(m.saturazione_stoccaggio_pct) as sat
        FROM monitoraggio m JOIN assets a ON a.id=m.asset_id
        WHERE a.tipo='magazzino' AND m.saturazione_stoccaggio_pct IS NOT NULL
    """).fetchone()
    s_mag = get_soglie('magazzino', 'saturazione_stoccaggio_pct')
    sat_pct = round(mag["sat"] or 0, 1)

    # KPI depositi: mezzi disponibili / totali (percentuale)
    dep = db.execute("""
        SELECT SUM(m.mezzi_disponibili) as disp, SUM(m.mezzi_totali) as tot
        FROM monitoraggio m JOIN assets a ON a.id=m.asset_id
        WHERE a.tipo='deposito'
    """).fetchone()
    s_dep = get_soglie('deposito', 'mezzi_disponibili')
    dep_disp = int(dep["disp"] or 0)
    dep_tot  = int(dep["tot"] or 0)
    dep_pct  = round((dep_disp / dep_tot * 100) if dep_tot > 0 else 0, 1)

    # Allarmi attivi
    allarmi_attivi = db.execute(
        "SELECT COUNT(*) FROM alarms WHERE ack_at IS NULL"
    ).fetchone()[0]
    allarmi_alarm = db.execute(
        "SELECT COUNT(*) FROM alarms WHERE livello='alarm' AND ack_at IS NULL"
    ).fetchone()[0]

    # Asset da monitorare (con allarmi attivi)
    da_monitorare = db.execute("""
        SELECT DISTINCT a.id, a.nome, a.tipo, a.citta,
               MAX(al.livello) as livello_max
        FROM alarms al JOIN assets a ON a.id=al.asset_id
        WHERE al.ack_at IS NULL
        GROUP BY a.id ORDER BY livello_max DESC LIMIT 5
    """).fetchall()

    return {
        "totale_asset": totale,
        "per_tipo": per_tipo,
        "kpi": {
            "uffici": {
                "label": "Occupancy uffici",
                "valore": occ_pct,
                "unita": "%",
                **s_uff
            },
            "stabilimenti": {
                "label": "Linee produzione attive",
                "valore": stab_att,
                "totale": stab_tot,
                "pct": stab_pct,
                "unita": "linee",
                **s_stab
            },
            "magazzini": {
                "label": "Saturazione stoccaggio",
                "valore": sat_pct,
                "unita": "%",
                **s_mag
            },
            "depositi": {
                "label": "Mezzi disponibili",
                "valore": dep_disp,
                "totale": dep_tot,
                "pct": dep_pct,
                "unita": "mezzi",
                **s_dep
            }
        },
        "allarmi": {
            "totale_attivi": allarmi_attivi,
            "livello_alarm": allarmi_alarm
        },
        "da_monitorare": [dict(r) for r in da_monitorare],
        "aggiornato_il": datetime.now().strftime("%d/%m/%Y %H:%M")
    }




# ── Funzione interna: ricalcolo allarmi ─────────────────────────────────────
def _ricalcola_allarmi_db(db):
    """
    Ricalcola tutti gli allarmi usando le soglie percentuali.
    - is_percentage=1 + campo_totale: pct = valore/totale*100
    - is_percentage=1 + campo_totale=None: il campo è già una percentuale
    - inverso=1: basso = brutto (alarm se pct < alarm_value)
    - inverso=0: alto = brutto (alarm se pct >= alarm_value)
    Il valore salvato in alarms.valore è sempre la percentuale calcolata.
    """
    db.execute("DELETE FROM alarms WHERE ack_at IS NULL")
    thresholds = db.execute("SELECT * FROM thresholds").fetchall()
    now = datetime.now().isoformat()
    for t in thresholds:
        t = dict(t)
        assets = db.execute(
            "SELECT id FROM assets WHERE tipo=?", (t["asset_tipo"],)
        ).fetchall()
        for asset in assets:
            mon = db.execute(
                "SELECT * FROM monitoraggio WHERE asset_id=?", (asset["id"],)
            ).fetchone()
            if not mon:
                continue
            mon = dict(mon)
            try:
                valore = mon.get(t["campo"])
                if valore is None:
                    continue
                valore = float(valore)
            except (KeyError, TypeError, ValueError):
                continue
            # Calcola percentuale
            if t["is_percentage"] and t["campo_totale"]:
                totale = mon.get(t["campo_totale"])
                if not totale or float(totale) == 0:
                    continue
                pct = (valore / float(totale)) * 100.0
            elif t["is_percentage"]:
                pct = valore
            else:
                pct = valore
            # Determina livello
            livello = None
            if t.get("inverso"):
                if pct < float(t["alarm_value"]):
                    livello = "alarm"
                elif pct < float(t["warning_value"]):
                    livello = "warning"
            else:
                if pct >= float(t["alarm_value"]):
                    livello = "alarm"
                elif pct >= float(t["warning_value"]):
                    livello = "warning"
            if livello:
                db.execute("""
                    INSERT INTO alarms (asset_id, campo, valore, livello, stato, created_at)
                    VALUES (?,?,?,?,?,?)
                """, (asset["id"], t["campo"], round(pct, 1), livello, livello, now))
    db.commit()

# ── Endpoint soglie ──────────────────────────────────────────────────────────
@app.get("/api/thresholds")
def lista_soglie(db=Depends(get_db), _=Depends(get_utente_corrente)):
    rows = db.execute("SELECT * FROM thresholds ORDER BY asset_tipo, campo").fetchall()
    return [dict(r) for r in rows]


@app.put("/api/thresholds/{threshold_id}")
def modifica_soglia(threshold_id: int, payload: ThresholdUpdate,
                    db=Depends(get_db), _=Depends(richiedi_admin)):
    soglia = db.execute("SELECT * FROM thresholds WHERE id=?", (threshold_id,)).fetchone()
    if not soglia:
        raise HTTPException(status_code=404, detail="Soglia non trovata")
    # Per soglie inverse (basso = brutto) warning_value > alarm_value: non bloccare
    is_inv = payload.inverso if payload.inverso is not None else soglia["inverso"]
    if not is_inv and payload.warning_value >= payload.alarm_value:
        raise HTTPException(status_code=400, detail="Il valore warning deve essere inferiore all'alarm")
    if is_inv and payload.warning_value <= payload.alarm_value:
        raise HTTPException(status_code=400, detail="Per soglie inverse il warning deve essere superiore all'alarm")
    # Aggiorna solo i campi forniti
    is_pct = payload.is_percentage if payload.is_percentage is not None else soglia["is_percentage"]
    ct = payload.campo_totale if payload.campo_totale is not None else soglia["campo_totale"]
    inv = is_inv
    db.execute(
        "UPDATE thresholds SET warning_value=?, alarm_value=?, is_percentage=?, campo_totale=?, inverso=? WHERE id=?",
        (payload.warning_value, payload.alarm_value, is_pct, ct, inv, threshold_id)
    )
    db.commit()
    # Ricalcola allarmi dopo modifica soglia
    _ricalcola_allarmi_db(db)
    return {"messaggio": "Soglia aggiornata e allarmi ricalcolati"}


# ── Endpoint allarmi ─────────────────────────────────────────────────────────
@app.get("/api/alarms")
def lista_allarmi(db=Depends(get_db), _=Depends(get_utente_corrente)):
    rows = db.execute("""
        SELECT al.*, a.nome as asset_nome, a.tipo as asset_tipo, a.citta
        FROM alarms al JOIN assets a ON a.id=al.asset_id
        ORDER BY al.livello DESC, al.created_at DESC
    """).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/alarms/{alarm_id}/ack")
def acknowledge_allarme(alarm_id: int, db=Depends(get_db), _=Depends(richiedi_admin)):
    allarme = db.execute("SELECT id FROM alarms WHERE id=?", (alarm_id,)).fetchone()
    if not allarme:
        raise HTTPException(status_code=404, detail="Allarme non trovato")
    db.execute(
        "UPDATE alarms SET ack_at=? WHERE id=?",
        (datetime.now().isoformat(), alarm_id)
    )
    db.commit()
    return {"messaggio": "Allarme confermato"}


# ── Endpoint configurazione ──────────────────────────────────────────────────
@app.get("/api/config")
def leggi_config(db=Depends(get_db), _=Depends(get_utente_corrente)):
    rows = db.execute("SELECT chiave, valore FROM config").fetchall()
    return {r["chiave"]: r["valore"] for r in rows}


@app.put("/api/config")
def aggiorna_config(payload: ConfigUpdate, db=Depends(get_db), _=Depends(richiedi_admin)):
    campi = {k: v for k, v in payload.dict().items() if v is not None}
    for chiave, valore in campi.items():
        db.execute(
            "INSERT OR REPLACE INTO config (chiave, valore) VALUES (?,?)",
            (chiave, valore)
        )
    db.commit()
    return {"messaggio": "Configurazione aggiornata"}


# ── Endpoint import Excel ────────────────────────────────────────────────────
@app.post("/api/assets/import")
async def importa_excel(file: UploadFile = File(...),
                        db=Depends(get_db), _=Depends(richiedi_admin)):
    """
    Importa asset da file Excel (.xlsx).
    Colonne attese: codice, nome, tipo, indirizzo, citta, provincia, cap,
                    lat, lon, referente, telefono, email,
                    superficie_mq, anno_costruzione, stato, note
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installato")

    contenuto = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenuto))
    ws = wb.active

    intestazioni = [cell.value for cell in ws[1]]
    campi_attesi = ["codice","nome","tipo","indirizzo","citta","provincia","cap",
                    "lat","lon","referente","telefono","email",
                    "superficie_mq","anno_costruzione","stato","note"]

    inseriti = 0
    aggiornati = 0
    errori = []

    for i, riga in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(riga):
            continue
        dati = dict(zip(intestazioni, riga))
        try:
            codice = str(dati.get("codice", "")).strip()
            if not codice:
                errori.append(f"Riga {i}: codice mancante")
                continue
            esistente = db.execute(
                "SELECT id FROM assets WHERE codice=?", (codice,)
            ).fetchone()
            if esistente:
                db.execute("""
                    UPDATE assets SET nome=?,tipo=?,indirizzo=?,citta=?,provincia=?,cap=?,
                    lat=?,lon=?,referente=?,telefono=?,email=?,superficie_mq=?,
                    anno_costruzione=?,stato=?,note=? WHERE codice=?
                """, (dati.get("nome"), dati.get("tipo"), dati.get("indirizzo"),
                      dati.get("citta"), dati.get("provincia"), dati.get("cap"),
                      dati.get("lat"), dati.get("lon"), dati.get("referente"),
                      dati.get("telefono"), dati.get("email"), dati.get("superficie_mq"),
                      dati.get("anno_costruzione"), dati.get("stato","attivo"),
                      dati.get("note"), codice))
                aggiornati += 1
            else:
                db.execute("""
                    INSERT INTO assets
                    (codice,nome,tipo,indirizzo,citta,provincia,cap,lat,lon,
                     referente,telefono,email,superficie_mq,anno_costruzione,stato,note)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (codice, dati.get("nome"), dati.get("tipo"), dati.get("indirizzo"),
                      dati.get("citta"), dati.get("provincia"), dati.get("cap"),
                      dati.get("lat"), dati.get("lon"), dati.get("referente"),
                      dati.get("telefono"), dati.get("email"), dati.get("superficie_mq"),
                      dati.get("anno_costruzione"), dati.get("stato","attivo"),
                      dati.get("note")))
                inseriti += 1
        except Exception as e:
            errori.append(f"Riga {i}: {str(e)}")

    db.commit()
    return {
        "inseriti": inseriti,
        "aggiornati": aggiornati,
        "errori": errori
    }


# ── Simulatore dati live ────────────────────────────────────────────────────
# Aggiorna i dati di monitoraggio ogni 60 secondi con variazioni graduali.
# Regole:
# - Uffici, magazzini, depositi: attivi 08:00-19:00 (fuori orario valori calano)
# - Stabilimenti: h24, nessuna riduzione notturna
# - Variazione graduale: ±5-8% per tick, con clamp ai limiti fisici

def _ora_lavorativa() -> bool:
    """Restituisce True se l'ora corrente è nell'orario lavorativo (08-19)."""
    ora = datetime.now().hour
    return 8 <= ora < 19


def _varia(valore: int, minimo: int, massimo: int,
           step_pct: float = 0.07, verso_min: bool = False) -> int:
    """
    Applica una variazione graduale al valore.
    - step_pct: ampiezza massima della variazione come % del massimo
    - verso_min: se True, spinge il valore verso il minimo (fuori orario)
    """
    if massimo <= 0:
        return valore
    step = max(1, int(massimo * step_pct))
    if verso_min:
        # Fuori orario: tende verso il 10% del massimo
        target = max(minimo, int(massimo * 0.10))
        delta = random.randint(0, step)
        nuovo = valore - delta if valore > target else valore + random.randint(0, 1)
    else:
        delta = random.randint(-step, step)
        nuovo = valore + delta
    return max(minimo, min(massimo, nuovo))


async def _simulatore_loop():
    """Task asincrono che aggiorna i dati di monitoraggio ogni 60 secondi."""
    await asyncio.sleep(10)  # attende l'avvio completo del server
    while True:
        try:
            conn = sqlite3.connect(DB_PATH, check_same_thread=False)
            conn.row_factory = sqlite3.Row
            now = datetime.now().strftime("%Y-%m-%d %H:%M")
            orario = _ora_lavorativa()

            assets = conn.execute(
                "SELECT a.id, a.tipo FROM assets a WHERE a.stato='attivo'"
            ).fetchall()

            for asset in assets:
                aid  = asset["id"]
                tipo = asset["tipo"]
                mon  = conn.execute(
                    "SELECT * FROM monitoraggio WHERE asset_id=?", (aid,)
                ).fetchone()
                if not mon:
                    continue
                mon = dict(mon)

                # Stabilimenti: h24, nessuna riduzione notturna
                if tipo == "stabilimento":
                    att  = mon.get("personale_attivo") or 0
                    cap  = mon.get("capacita_personale") or 0
                    latt = mon.get("linee_produzione_attive") or 0
                    ltot = mon.get("linee_produzione_totali") or 0
                    if cap > 0:
                        att  = _varia(att,  0, cap)
                    if ltot > 0:
                        latt = _varia(latt, 0, ltot)
                    conn.execute(
                        "UPDATE monitoraggio SET aggiornato_il=?, personale_attivo=?,"
                        " linee_produzione_attive=? WHERE asset_id=?",
                        (now, att, latt, aid)
                    )

                # Uffici: occupancy e sale riunioni
                elif tipo == "ufficio":
                    dip  = mon.get("dipendenti_presenti") or 0
                    cap  = mon.get("capienza_massima") or 0
                    socc = mon.get("sale_riunioni_occupate") or 0
                    stot = mon.get("sale_riunioni_totali") or 0
                    if cap > 0:
                        dip  = _varia(dip,  0, cap,  verso_min=not orario)
                    if stot > 0:
                        socc = _varia(socc, 0, stot, verso_min=not orario)
                    conn.execute(
                        "UPDATE monitoraggio SET aggiornato_il=?, dipendenti_presenti=?,"
                        " sale_riunioni_occupate=? WHERE asset_id=?",
                        (now, dip, socc, aid)
                    )

                # Magazzini: saturazione e mezzi operativi
                elif tipo == "magazzino":
                    sat  = mon.get("saturazione_stoccaggio_pct") or 0
                    mop  = mon.get("mezzi_magazzino_presenti") or 0
                    mtot = mon.get("mezzi_magazzino_totali") or 0
                    sat  = _varia(sat,  0, 100, step_pct=0.05, verso_min=not orario)
                    if mtot > 0:
                        mop = _varia(mop, 0, mtot, verso_min=not orario)
                    conn.execute(
                        "UPDATE monitoraggio SET aggiornato_il=?, saturazione_stoccaggio_pct=?,"
                        " mezzi_magazzino_presenti=? WHERE asset_id=?",
                        (now, sat, mop, aid)
                    )

                # Depositi: mezzi disponibili, in missione, in manutenzione
                elif tipo == "deposito":
                    mtot = mon.get("mezzi_totali") or 0
                    mman = mon.get("mezzi_in_manutenzione") or 0
                    mmis = mon.get("mezzi_in_missione") or 0
                    if mtot > 0:
                        mman = _varia(mman, 0, max(1, mtot // 4), verso_min=not orario)
                        mmis = _varia(mmis, 0, max(1, mtot // 2), verso_min=not orario)
                        mdisp = max(0, mtot - mman - mmis)
                        conn.execute(
                            "UPDATE monitoraggio SET aggiornato_il=?, mezzi_in_manutenzione=?,"
                            " mezzi_in_missione=?, mezzi_disponibili=? WHERE asset_id=?",
                            (now, mman, mmis, mdisp, aid)
                        )

            conn.commit()

            # Ricalcola allarmi dopo ogni aggiornamento
            _ricalcola_allarmi_db(conn)
            conn.close()

        except Exception as exc:
            # Non blocca il loop in caso di errore transitorio
            print(f"[simulatore] errore: {exc}")

        await asyncio.sleep(60)


@app.on_event("startup")
async def avvia_simulatore():
    """Avvia il task di simulazione dati live all'avvio del server."""
    asyncio.create_task(_simulatore_loop())
    print("[simulatore] task live avviato (intervallo: 60s)")


# ── Endpoint ricalcolo manuale allarmi ───────────────────────────────────────
@app.post("/api/alarms/recalculate")
def ricalcola_manuale(db=Depends(get_db), _=Depends(richiedi_admin)):
    """Forza il ricalcolo immediato degli allarmi (utile per test)."""
    _ricalcola_allarmi_db(db)
    n = db.execute("SELECT COUNT(*) FROM alarms WHERE ack_at IS NULL").fetchone()[0]
    return {"messaggio": "Ricalcolo completato", "allarmi_attivi": n}


# ── Redirect root → login ────────────────────────────────────────────────────
@app.get("/")
def root():
    return RedirectResponse(url="/static/login.html")


# ── Serve file statici ───────────────────────────────────────────────────────
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
